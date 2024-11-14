from openpyxl import load_workbook

def get_datas_from_xlsx(filename: str, sheet_page: int) -> list[tuple]:
    """
    Из xlsx файла прочитает и возвращает список с материалами 'list[tuple]'

    :param filename: Имя файла
    :param sheet_page: Индекс страницы
    :return: Список с материалами
    """
    # На этот метод можно добавить валидацию листа, пустых строк, ожидаемые строки и т.д.
    workbook = load_workbook(filename=filename)
    sheet = workbook.worksheets[sheet_page]

    result = []
    for rows in sheet.iter_rows(min_row=2, values_only=True):
        result.append(rows)

    return result